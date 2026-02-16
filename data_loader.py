import io
import pandas as pd
import requests
from openpyxl import load_workbook


def download_xlsx_to_bytes(xlsx_url):
    resp = requests.get(xlsx_url, timeout=120)
    resp.raise_for_status()
    return resp.content


def _load_all_sheets_as_tables_from_bytes(xlsx_bytes, selected_sheets=None):
    bio = io.BytesIO(xlsx_bytes)
    wb = load_workbook(bio, data_only=True)

    sheet_names = wb.sheetnames
    if selected_sheets is not None and len(selected_sheets) > 0:
        sheet_names = [s for s in sheet_names if s in selected_sheets]

    tables = {}
    meta_rows = []

    for sheet_name in sheet_names:
        bio_sheet = io.BytesIO(xlsx_bytes)
        df = pd.read_excel(bio_sheet, sheet_name=sheet_name)

        key = "sheet::" + str(sheet_name)
        tables[key] = df

        meta_rows.append(
            {
                "key": key,
                "sheet_name": str(sheet_name),
                "n_rows": int(df.shape[0]),
                "n_cols": int(df.shape[1]),
            }
        )

    meta_df = pd.DataFrame(meta_rows)
    return tables, meta_df, wb.sheetnames


def load_workbook_tables_from_url(xlsx_url, selected_sheets=None, min_text_cells=4):
    xlsx_bytes = download_xlsx_to_bytes(xlsx_url)

    tables_all, meta_all, all_sheets = _load_all_sheets_as_tables_from_bytes(
        xlsx_bytes, selected_sheets=selected_sheets
    )

    tables_filtered = {}
    meta_rows = []

    for key in list(tables_all.keys()):
        df = tables_all[key]

        if df is None:
            continue

        if df.shape[0] == 0 or df.shape[1] == 0:
            continue

        df_head = df.head(25)

        non_null_cells = int(df_head.notna().sum().sum())
        if non_null_cells < 20:
            continue

        text_like_cells = 0
        try:
            obj_df = df_head.select_dtypes(include=["object"])
            text_like_cells = int(obj_df.notna().sum().sum())
        except Exception:
            text_like_cells = 0

        if text_like_cells < int(min_text_cells):
            continue

        tables_filtered[key] = df

        sheet_name_val = ""
        if key.startswith("sheet::"):
            sheet_name_val = key.replace("sheet::", "")

        meta_rows.append(
            {
                "key": key,
                "sheet_name": sheet_name_val,
                "n_rows": int(df.shape[0]),
                "n_cols": int(df.shape[1]),
                "non_null_cells_head25": non_null_cells,
                "text_like_cells_head25": text_like_cells,
            }
        )

    meta_df = pd.DataFrame(meta_rows)

    if len(tables_filtered) == 0:
        meta_all_copy = meta_all.copy()
        meta_all_copy["note"] = "fallback_all_sheets"
        return tables_all, meta_all_copy, all_sheets

    return tables_filtered, meta_df, all_sheets
